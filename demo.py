import sys
import os

import requests
import time

from openpyxl.reader.excel import load_workbook

from PyQt5 import QtCore
from PyQt5.Qt import QThread
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QMainWindow, QApplication, QDesktopWidget, QVBoxLayout, QLabel

from main1017 import Ui_MainWindow

# from main1017_temp1 import Ui_TempWindow
# from main1017_hum import Ui_HumWindow
# from main1017_lum import Ui_LumWindow
# from main1017_vol import Ui_VolWindow
# from main1017_mag import Ui_MagWindow


WIN_WIDTH = 1088
WIN_HEIGHT = 670

FILENAME_temp = "tem_value_stream.xlsx"
FILENAME_hum = "hum_value_stream.xlsx"
FILENAME_lum = "lum_value_stream.xlsx"
FILENAME_vol = "vol_value_stream.xlsx"
FILENAME_mag = "mag_value_stream.xlsx"


class TempThread(QThread):

    def __init__(self):
        super().__init__()
        self.load_my_exel()

        self.max = -273.0
        self.min = 1000.0
        self.flag = True
        self.timeStamp_s = 0
        self.timeStamp_ms = 0
        self.offline_counter = 0
        self.state = None

    def run(self):
        while True:
            time.sleep(0.45)
            response = requests.get(
                url="https://iot-api.heclouds.com/thingmodel/query-device-property?product_id=sF3UI8miN9&device_name=dht_11",
                headers={"Accept": "application/json, text/plain, */*",
                         "authorization": "version=2020-05-29&res=userid%2F345882&et=1720540119&method=sha1&sign=VLZuKbgiN5XqfgbexUodz9djZAg%3D"})

            if response.status_code == 200:
                pass
                if self.timeStamp_ms != 0:
                    if self.timeStamp_ms == int(response.json()['data'][0]['time']):
                        self.flag = False
                        self.offline_counter += 1
                        if self.offline_counter >= 3:  # 控制上下线检测延时 todo
                            if w_main.temp_state.text() == "在线":
                                localtime = time.localtime()
                                output_time = str(localtime[0]) + '/' + str(localtime[1]) + '/' + str(
                                    localtime[2]) + '  ' + str(localtime[3]) + ':' + str(localtime[4]) + ':' + str(
                                    localtime[5])
                                file = open("tem_log.txt", "a")
                                file.write(output_time + "\t离线\n")
                                file.close()

                                msg_history = []
                                file = open("tem_log.txt", "r")
                                for _ in range(50):
                                    log = file.readline()
                                    msg_history.append(log)
                                    if log == "":
                                        file.close()
                                        break
                                w_main.msg_temp.setText(''.join(msg_history))
                                w_main.msg_temp.resize(440, w_main.msg_temp.frameSize().height() + 15)
                            w_main.temp_state.setText("离线")
                            w_main.home_temp_state.setStyleSheet(
                                "image: url(:/loo/unconnect.png);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                            w_main.frame_state_temp.setStyleSheet(
                                "image: url(:/loo/unconnect.png);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                            self.state = False

                    else:
                        if w_main.temp_state.text() == "离线":
                            localtime = time.localtime()
                            output_time = str(localtime[0]) + '/' + str(localtime[1]) + '/' + str(
                                localtime[2]) + '  ' + str(localtime[3]) + ':' + str(localtime[4]) + ':' + str(
                                localtime[5])
                            # print(localtime[0], type(localtime[0]))
                            file = open("tem_log.txt", "a")
                            file.write(output_time + "\t上线\n")
                            file.close()

                            msg_history = []
                            file = open("tem_log.txt", "r")
                            for _ in range(50):
                                log = file.readline()
                                msg_history.append(log)
                                if log == "":
                                    file.close()
                                    break
                            w_main.msg_temp.setText(''.join(msg_history))
                            w_main.msg_temp.resize(440, w_main.msg_temp.frameSize().height() + 15)
                        self.flag = True
                        self.offline_counter = 0
                        w_main.temp_state.setText("在线")  # todo
                        w_main.home_temp_state.setStyleSheet(
                            "image: url(:/loo/connect.jpg);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                        w_main.frame_state_temp.setStyleSheet(
                            "image: url(:/loo/connect.jpg);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                        self.state = True

                self.timeStamp_ms = int(response.json()['data'][0]['time'])

                # 获取数据并处理
                tem_value = response.json()['data'][1]['value']
                # print(type(self.max), float(tem_value))
                if self.max <= float(tem_value):
                    self.max = float(tem_value)
                    w_main.temp_max.setText(tem_value)
                if self.min >= float(tem_value):
                    self.min = float(tem_value)
                    w_main.temp_min.setText(tem_value)

                if self.state == True:
                    w_main.home_temp_value.setText("  " + tem_value)
                    w_main.temp_temp_value.setText(tem_value)
                else:
                    w_main.home_temp_value.setText("————")
                    w_main.temp_temp_value.setText("————")
                # print(tem_value)

            else:
                print('Failed to get data, status code:', response.status_code)
                break

            ms = self.timeStamp_ms % 1000
            self.timeStamp_s = self.timeStamp_ms // 1000

            if self.flag:
                self.timeArray = time.localtime(self.timeStamp_s)

                self.timearray = []

                for i in range(6):
                    self.timearray.append(str(self.timeArray[i]))
                self.timearray.append(str(ms))

                output_time = self.timearray[0] + '/' + self.timearray[1] + '/' + self.timearray[2] + '  ' + \
                              self.timearray[3] + ':' + self.timearray[4] + ':' + self.timearray[5] + '.' + \
                              self.timearray[6]

                # 输入获取的数据
                self.ws.cell(row=self.r, column=1).value = output_time
                # print(output_time)
                self.ws.cell(row=self.r, column=2).value = tem_value
                self.r += 1

                self.r_temp = self.r - 1
                self.ws.cell(row=1, column=5).value = self.r_temp
                self.wb.save(FILENAME_temp)

    def load_my_exel(self):
        if os.path.exists(FILENAME_temp):
            # 选中对应工作表
            self.wb = load_workbook(FILENAME_temp)
            self.ws = self.wb.active

            # 设置标题行
            self.ws.cell(row=1, column=1).value = '时间'
            self.ws.cell(row=1, column=2).value = '温度/°C'
            self.ws.cell(row=1, column=4).value = '行数:'

            # 读取历史数据量
            if self.ws.cell(row=1, column=5).value is None or self.ws.cell(row=1, column=5).value <= 0:
                self.ws.cell(row=1, column=5).value = 1
            self.r = self.ws.cell(row=1, column=5).value + 1

            # 设置列宽
            self.ws.column_dimensions['A'].width = 30
            self.ws.column_dimensions['B'].width = 10
            self.ws.column_dimensions['C'].width = 10
            self.ws.column_dimensions['D'].width = 10
            self.ws.column_dimensions['E'].width = 10


class HumThread(QThread):

    def __init__(self):
        super().__init__()
        self.load_my_exel()

        self.max = -1.0
        self.min = 200.0
        self.flag = True
        self.timeStamp_s = 0
        self.timeStamp_ms = 0
        self.offline_counter = 0
        self.state = None

    def run(self):
        while True:
            time.sleep(0.45)
            response = requests.get(
                url="https://iot-api.heclouds.com/thingmodel/query-device-property?product_id=sF3UI8miN9&device_name=dht_11",
                headers={"Accept": "application/json, text/plain, */*",
                         "authorization": "version=2020-05-29&res=userid%2F345882&et=1720540119&method=sha1&sign=VLZuKbgiN5XqfgbexUodz9djZAg%3D"})

            if response.status_code == 200:
                if self.timeStamp_ms != 0:
                    if self.timeStamp_ms == int(response.json()['data'][0]['time']):
                        self.flag = False
                        self.offline_counter += 1
                        if self.offline_counter >= 3:  # 控制上下线检测延时 tobedone
                            if w_main.hum_state.text() == "在线":
                                localtime = time.localtime()
                                output_time = str(localtime[0]) + '/' + str(localtime[1]) + '/' + str(
                                    localtime[2]) + '  ' + str(localtime[3]) + ':' + str(localtime[4]) + ':' + str(
                                    localtime[5])
                                file = open("hum_log.txt", "a")
                                file.write(output_time + "\t离线\n")
                                file.close()

                                msg_history = []
                                file = open("hum_log.txt", "r")
                                for _ in range(50):
                                    log = file.readline()
                                    msg_history.append(log)
                                    if log == "":
                                        file.close()
                                        break
                                w_main.msg_hum.setText(''.join(msg_history))
                                w_main.msg_hum.resize(440, w_main.msg_hum.frameSize().height() + 15)
                            w_main.hum_state.setText("离线")
                            w_main.state_main_hum.setStyleSheet(
                                "image: url(:/loo/unconnect.png);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                            w_main.frame_state_hum.setStyleSheet(
                                "image: url(:/loo/unconnect.png);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                            self.state = False

                    else:
                        if w_main.hum_state.text() == "离线":
                            localtime = time.localtime()
                            output_time = str(localtime[0]) + '/' + str(localtime[1]) + '/' + str(
                                localtime[2]) + '  ' + str(localtime[3]) + ':' + str(localtime[4]) + ':' + str(
                                localtime[5])
                            # print(localtime[0], type(localtime[0]))
                            file = open("hum_log.txt", "a")
                            file.write(output_time + "\t上线\n")
                            file.close()

                            msg_history = []
                            file = open("hum_log.txt", "r")
                            for _ in range(50):
                                log = file.readline()
                                msg_history.append(log)
                                if log == "":
                                    file.close()
                                    break
                            w_main.msg_hum.setText(''.join(msg_history))
                            w_main.msg_hum.resize(440, w_main.msg_hum.frameSize().height() + 15)
                        self.flag = True
                        self.offline_counter = 0
                        w_main.hum_state.setText("在线")  # todo
                        w_main.state_main_hum.setStyleSheet(
                            "image: url(:/loo/connect.jpg);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                        w_main.frame_state_hum.setStyleSheet(
                            "image: url(:/loo/connect.jpg);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                        self.state = True

                self.timeStamp_ms = int(response.json()['data'][0]['time'])

                # 获取数据并处理
                hum_value = response.json()['data'][0]['value']
                # print(type(self.max), float(hum_value))
                if self.max <= float(hum_value):
                    self.max = float(hum_value)
                    w_main.hum_max.setText(hum_value)
                if self.min >= float(hum_value):
                    self.min = float(hum_value)
                    w_main.hum_min.setText(hum_value)

                if self.state == True:
                    w_main.main_hum_value.setText("  " + hum_value)
                    w_main.hum_hum_value.setText(hum_value)
                else:
                    w_main.main_hum_value.setText("————")
                    w_main.hum_hum_value.setText("————")
                # print(hum_value)

            else:
                print('Failed to get data, status code:', response.status_code)
                break

            ms = self.timeStamp_ms % 1000
            self.timeStamp_s = self.timeStamp_ms // 1000

            if self.flag:
                self.timeArray = time.localtime(self.timeStamp_s)

                self.timearray = []

                for i in range(6):
                    self.timearray.append(str(self.timeArray[i]))
                self.timearray.append(str(ms))

                output_time = self.timearray[0] + '/' + self.timearray[1] + '/' + self.timearray[2] + '  ' + \
                              self.timearray[3] + ':' + self.timearray[4] + ':' + self.timearray[5] + '.' + \
                              self.timearray[6]

                # 输入获取的数据
                self.ws.cell(row=self.r, column=1).value = output_time
                # print(output_time)
                self.ws.cell(row=self.r, column=2).value = hum_value
                self.r += 1

                self.r_temp = self.r - 1
                self.ws.cell(row=1, column=5).value = self.r_temp
                self.wb.save(FILENAME_hum)

    def load_my_exel(self):
        if os.path.exists(FILENAME_hum):
            # 选中对应工作表
            self.wb = load_workbook(FILENAME_hum)
            self.ws = self.wb.active

            # 设置标题行
            self.ws.cell(row=1, column=1).value = '时间'
            self.ws.cell(row=1, column=2).value = '湿度/%'
            self.ws.cell(row=1, column=4).value = '行数:'

            # 读取历史数据量
            if self.ws.cell(row=1, column=5).value is None or self.ws.cell(row=1, column=5).value <= 0:
                self.ws.cell(row=1, column=5).value = 1
            self.r = self.ws.cell(row=1, column=5).value + 1

            # 设置列宽
            self.ws.column_dimensions['A'].width = 30
            self.ws.column_dimensions['B'].width = 10
            self.ws.column_dimensions['C'].width = 10
            self.ws.column_dimensions['D'].width = 10
            self.ws.column_dimensions['E'].width = 10


class LumThread(QThread):

    def __init__(self):
        super().__init__()
        self.load_my_exel()

        self.max = -1.0
        self.min = 100000.0
        self.flag = True
        self.timeStamp_s = 0
        self.timeStamp_ms = 0
        self.offline_counter = 0
        self.state = None

    def run(self):
        while True:
            time.sleep(0.45)
            response = requests.get(
                url="https://iot-api.heclouds.com/thingmodel/query-device-property?product_id=6d4XVMbY0f&device_name=gy_30",
                headers={"Accept": "application/json, text/plain, */*",
                         "authorization": "version=2020-05-29&res=userid%2F345882&et=1720540119&method=sha1&sign=VLZuKbgiN5XqfgbexUodz9djZAg%3D"})

            if response.status_code == 200:
                if self.timeStamp_ms != 0:
                    if self.timeStamp_ms == int(response.json()['data'][0]['time']):
                        self.flag = False
                        self.offline_counter += 1
                        if self.offline_counter >= 3:  # 控制上下线检测延时 tobedone
                            if w_main.lum_state.text() == "在线":
                                localtime = time.localtime()
                                output_time = str(localtime[0]) + '/' + str(localtime[1]) + '/' + str(
                                    localtime[2]) + '  ' + str(localtime[3]) + ':' + str(localtime[4]) + ':' + str(
                                    localtime[5])
                                file = open("lum_log.txt", "a")
                                file.write(output_time + "\t离线\n")
                                file.close()

                                msg_history = []
                                file = open("lum_log.txt", "r")
                                for _ in range(50):
                                    log = file.readline()
                                    msg_history.append(log)
                                    if log == "":
                                        file.close()
                                        break
                                w_main.msg_lum.setText(''.join(msg_history))
                                w_main.msg_lum.resize(440, w_main.msg_lum.frameSize().height() + 15)
                            w_main.lum_state.setText("离线")
                            w_main.state_main_lum.setStyleSheet(
                                "image: url(:/loo/unconnect.png);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                            w_main.frame_state_lum.setStyleSheet(
                                "image: url(:/loo/unconnect.png);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                            self.state = False

                    else:
                        if w_main.lum_state.text() == "离线":
                            localtime = time.localtime()
                            output_time = str(localtime[0]) + '/' + str(localtime[1]) + '/' + str(
                                localtime[2]) + '  ' + str(localtime[3]) + ':' + str(localtime[4]) + ':' + str(
                                localtime[5])
                            # print(localtime[0], type(localtime[0]))
                            file = open("lum_log.txt", "a")
                            file.write(output_time + "\t上线\n")
                            file.close()

                            msg_history = []
                            file = open("lum_log.txt", "r")
                            for _ in range(50):
                                log = file.readline()
                                msg_history.append(log)
                                if log == "":
                                    file.close()
                                    break
                            w_main.msg_lum.setText(''.join(msg_history))
                            w_main.msg_lum.resize(440, w_main.msg_lum.frameSize().height() + 15)
                        self.flag = True
                        self.offline_counter = 0
                        w_main.lum_state.setText("在线")  # todo
                        w_main.state_main_lum.setStyleSheet(
                            "image: url(:/loo/connect.jpg);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                        w_main.frame_state_lum.setStyleSheet(
                            "image: url(:/loo/connect.jpg);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                        self.state = True

                self.timeStamp_ms = int(response.json()['data'][0]['time'])

                # 获取数据并处理
                lum_value = response.json()['data'][0]['value']
                # print(type(self.max), float(hum_value))
                if self.max <= float(lum_value):
                    self.max = float(lum_value)
                    w_main.lum_max.setText(lum_value)
                if self.min >= float(lum_value):
                    self.min = float(lum_value)
                    w_main.lum_min.setText(lum_value)

                if self.state == True:
                    w_main.main_lum_value.setText("  " + lum_value)
                    w_main.lum_lum_value.setText(lum_value)
                else:
                    w_main.main_lum_value.setText("————")
                    w_main.lum_lum_value.setText("————")
                # print(lum_value)

            else:
                # print('Failed to get data, status code:', response.status_code)
                break

            ms = self.timeStamp_ms % 1000
            self.timeStamp_s = self.timeStamp_ms // 1000

            if self.flag:
                self.timeArray = time.localtime(self.timeStamp_s)

                self.timearray = []

                for i in range(6):
                    self.timearray.append(str(self.timeArray[i]))
                self.timearray.append(str(ms))

                output_time = self.timearray[0] + '/' + self.timearray[1] + '/' + self.timearray[2] + '  ' + \
                              self.timearray[3] + ':' + self.timearray[4] + ':' + self.timearray[5] + '.' + \
                              self.timearray[6]

                # 输入获取的数据
                self.ws.cell(row=self.r, column=1).value = output_time
                # print(output_time)
                self.ws.cell(row=self.r, column=2).value = lum_value
                self.r += 1

                self.r_temp = self.r - 1
                self.ws.cell(row=1, column=5).value = self.r_temp
                self.wb.save(FILENAME_lum)

    def load_my_exel(self):
        if os.path.exists(FILENAME_lum):
            # 选中对应工作表
            self.wb = load_workbook(FILENAME_lum)
            self.ws = self.wb.active

            # 设置标题行
            self.ws.cell(row=1, column=1).value = '时间'
            self.ws.cell(row=1, column=2).value = '光强/lx'
            self.ws.cell(row=1, column=4).value = '行数:'

            # 读取历史数据量
            if self.ws.cell(row=1, column=5).value is None or self.ws.cell(row=1, column=5).value <= 0:
                self.ws.cell(row=1, column=5).value = 1
            self.r = self.ws.cell(row=1, column=5).value + 1

            # 设置列宽
            self.ws.column_dimensions['A'].width = 30
            self.ws.column_dimensions['B'].width = 10
            self.ws.column_dimensions['C'].width = 10
            self.ws.column_dimensions['D'].width = 10
            self.ws.column_dimensions['E'].width = 10


class VolThread(QThread):

    def __init__(self):
        super().__init__()
        self.load_my_exel()

        self.max = -1.0
        self.min = 200.0
        self.flag = True
        self.timeStamp_s = 0
        self.timeStamp_ms = 0
        self.offline_counter = 0
        self.state = None

    def run(self):
        while True:
            time.sleep(0.45)
            response = requests.get(
                url="https://iot-api.heclouds.com/thingmodel/query-device-property?product_id=ek40erZrGx&device_name=volume",
                headers={"Accept": "application/json, text/plain, */*",
                         "authorization": "version=2020-05-29&res=userid%2F345882&et=1720540119&method=sha1&sign=VLZuKbgiN5XqfgbexUodz9djZAg%3D"})

            if response.status_code == 200:
                if self.timeStamp_ms != 0:
                    if self.timeStamp_ms == int(response.json()['data'][0]['time']):
                        self.flag = False
                        self.offline_counter += 1
                        if self.offline_counter >= 3:  # 控制上下线检测延时 tobedone
                            if w_main.vol_state.text() == "在线":
                                localtime = time.localtime()
                                output_time = str(localtime[0]) + '/' + str(localtime[1]) + '/' + str(
                                    localtime[2]) + '  ' + str(localtime[3]) + ':' + str(localtime[4]) + ':' + str(
                                    localtime[5])
                                file = open("vol_log.txt", "a")
                                file.write(output_time + "\t离线\n")
                                file.close()

                                msg_history = []
                                file = open("vol_log.txt", "r")
                                for _ in range(50):
                                    log = file.readline()
                                    msg_history.append(log)
                                    if log == "":
                                        file.close()
                                        break
                                w_main.msg_vol.setText(''.join(msg_history))
                                w_main.msg_vol.resize(440, w_main.msg_vol.frameSize().height() + 15)
                            w_main.vol_state.setText("离线")
                            w_main.state_main_volume.setStyleSheet(
                                "image: url(:/loo/unconnect.png);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                            w_main.frame_state_vol.setStyleSheet(
                                "image: url(:/loo/unconnect.png);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                            self.state = False

                    else:
                        if w_main.vol_state.text() == "离线":
                            localtime = time.localtime()
                            output_time = str(localtime[0]) + '/' + str(localtime[1]) + '/' + str(
                                localtime[2]) + '  ' + str(localtime[3]) + ':' + str(localtime[4]) + ':' + str(
                                localtime[5])
                            # print(localtime[0], type(localtime[0]))
                            file = open("vol_log.txt", "a")
                            file.write(output_time + "\t上线\n")
                            file.close()

                            msg_history = []
                            file = open("vol_log.txt", "r")
                            for _ in range(50):
                                log = file.readline()
                                msg_history.append(log)
                                if log == "":
                                    file.close()
                                    break
                            w_main.msg_vol.setText(''.join(msg_history))
                            w_main.msg_vol.resize(440, w_main.msg_vol.frameSize().height() + 15)
                        self.flag = True
                        self.offline_counter = 0
                        w_main.vol_state.setText("在线")  # todo
                        w_main.state_main_volume.setStyleSheet(
                            "image: url(:/loo/connect.jpg);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                        w_main.frame_state_vol.setStyleSheet(
                            "image: url(:/loo/connect.jpg);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                        self.state = True

                self.timeStamp_ms = int(response.json()['data'][0]['time'])

                # 获取数据并处理
                vol_value = response.json()['data'][0]['value']
                # print(type(self.max), float(hum_value))
                if self.max <= float(vol_value):
                    self.max = float(vol_value)
                    w_main.vol_max.setText(vol_value)
                if self.min >= float(vol_value):
                    self.min = float(vol_value)
                    w_main.vol_min.setText(vol_value)

                if self.state == True:
                    w_main.volume_value.setText("  " + vol_value)
                    w_main.vol_vol_value.setText(vol_value)
                else:
                    w_main.volume_value.setText("————")
                    w_main.vol_vol_value.setText("————")
                # print(vol_value)

            else:
                print('Failed to get data, status code:', response.status_code)
                break

            ms = self.timeStamp_ms % 1000
            self.timeStamp_s = self.timeStamp_ms // 1000

            if self.flag:
                self.timeArray = time.localtime(self.timeStamp_s)

                self.timearray = []

                for i in range(6):
                    self.timearray.append(str(self.timeArray[i]))
                self.timearray.append(str(ms))

                output_time = self.timearray[0] + '/' + self.timearray[1] + '/' + self.timearray[2] + '  ' + \
                              self.timearray[3] + ':' + self.timearray[4] + ':' + self.timearray[5] + '.' + \
                              self.timearray[6]

                # 输入获取的数据
                self.ws.cell(row=self.r, column=1).value = output_time
                # print(output_time)
                self.ws.cell(row=self.r, column=2).value = vol_value
                self.r += 1

                self.r_temp = self.r - 1
                self.ws.cell(row=1, column=5).value = self.r_temp
                self.wb.save(FILENAME_vol)

    def load_my_exel(self):
        if os.path.exists(FILENAME_vol):
            # 选中对应工作表
            self.wb = load_workbook(FILENAME_vol)
            self.ws = self.wb.active

            # 设置标题行
            self.ws.cell(row=1, column=1).value = '时间'
            self.ws.cell(row=1, column=2).value = '声强系数'
            self.ws.cell(row=1, column=4).value = '行数:'

            # 读取历史数据量
            if self.ws.cell(row=1, column=5).value is None or self.ws.cell(row=1, column=5).value <= 0:
                self.ws.cell(row=1, column=5).value = 1
            self.r = self.ws.cell(row=1, column=5).value + 1

            # 设置列宽
            self.ws.column_dimensions['A'].width = 30
            self.ws.column_dimensions['B'].width = 10
            self.ws.column_dimensions['C'].width = 10
            self.ws.column_dimensions['D'].width = 10
            self.ws.column_dimensions['E'].width = 10


class MagThread(QThread):

    def __init__(self):
        super().__init__()
        self.load_my_exel()

        self.max = -273.0
        self.min = 1000.0
        self.flag = True
        self.timeStamp_s = 0
        self.timeStamp_ms = 0
        self.offline_counter = 0
        self.state = None

    def run(self):
        while True:
            time.sleep(0.45)
            response = requests.get(
                url="https://iot-api.heclouds.com/thingmodel/query-device-property?product_id=xe7J7J62KU&device_name=hall_3144",
                headers={"Accept": "application/json, text/plain, */*",
                         "authorization": "version=2020-05-29&res=userid%2F345882&et=1720540119&method=sha1&sign=VLZuKbgiN5XqfgbexUodz9djZAg%3D"})

            if response.status_code == 200:
                if self.timeStamp_ms != 0:
                    if self.timeStamp_ms == int(response.json()['data'][0]['time']):
                        self.flag = False
                        self.offline_counter += 1
                        if self.offline_counter >= 3:  # 控制上下线检测延时 tobedone
                            if w_main.mag_state.text() == "在线":
                                localtime = time.localtime()
                                output_time = str(localtime[0]) + '/' + str(localtime[1]) + '/' + str(
                                    localtime[2]) + '  ' + str(localtime[3]) + ':' + str(localtime[4]) + ':' + str(
                                    localtime[5])
                                file = open("mag_log.txt", "a")
                                file.write(output_time + "\t离线\n")
                                file.close()

                                msg_history = []
                                file = open("mag_log.txt", "r")
                                for _ in range(50):
                                    log = file.readline()
                                    msg_history.append(log)
                                    if log == "":
                                        file.close()
                                        break
                                w_main.msg_mag.setText(''.join(msg_history))
                                w_main.msg_mag.resize(440, w_main.msg_mag.frameSize().height() + 15)
                            w_main.mag_state.setText("离线")
                            w_main.state_main_mag.setStyleSheet(
                                "image: url(:/loo/unconnect.png);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                            w_main.frame_state_mag.setStyleSheet(
                                "image: url(:/loo/unconnect.png);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                            self.state = False

                    else:
                        self.flag = True
                        self.offline_counter = 0
                        w_main.label_10.setText("在线")  # todo
                        w_main.state_main_mag.setStyleSheet(
                            "image: url(:/loo/connect.jpg);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                        w_main.frame_state_mag.setStyleSheet(
                            "image: url(:/loo/connect.jpg);")  # image: url(:/loo/unconnect.png);  image: url(:/loo/connect.jpg);
                        self.state = True

                self.timeStamp_ms = int(response.json()['data'][0]['time'])

                # 获取数据并处理
                mag_value = response.json()['data'][1]['value']

                if self.state == True:
                    if mag_value == 1:
                        w_main.mag_value.setText(" 正常")
                        w_main.mag_mag_value.setText(" 正常")
                    else:
                        w_main.mag_value.setText(" 异常")
                        w_main.mag_mag_value.setText(" 异常")
                else:
                    w_main.mag_value.setText("————")
                    w_main.mag_mag_value.setText("————")

                # print(mag_value)
            else:
                print('Failed to get data, status code:', response.status_code)
                break

            ms = self.timeStamp_ms % 1000
            self.timeStamp_s = self.timeStamp_ms // 1000

            if self.flag:
                self.timeArray = time.localtime(self.timeStamp_s)

                self.timearray = []

                for i in range(6):
                    self.timearray.append(str(self.timeArray[i]))
                self.timearray.append(str(ms))

                output_time = self.timearray[0] + '/' + self.timearray[1] + '/' + self.timearray[2] + '  ' + \
                              self.timearray[3] + ':' + self.timearray[4] + ':' + self.timearray[5] + '.' + \
                              self.timearray[6]

                # 输入获取的数据
                self.ws.cell(row=self.r, column=1).value = output_time
                # print(output_time)
                self.ws.cell(row=self.r, column=2).value = mag_value
                self.r += 1

                self.r_temp = self.r - 1
                self.ws.cell(row=1, column=5).value = self.r_temp
                self.wb.save(FILENAME_mag)

    def load_my_exel(self):
        if os.path.exists(FILENAME_mag):
            # 选中对应工作表
            self.wb = load_workbook(FILENAME_mag)
            self.ws = self.wb.active

            # 设置标题行
            self.ws.cell(row=1, column=1).value = '时间'
            self.ws.cell(row=1, column=2).value = '磁场环境'
            self.ws.cell(row=1, column=4).value = '行数:'

            # 读取历史数据量
            if self.ws.cell(row=1, column=5).value is None or self.ws.cell(row=1, column=5).value <= 0:
                self.ws.cell(row=1, column=5).value = 1
            self.r = self.ws.cell(row=1, column=5).value + 1

            # 设置列宽
            self.ws.column_dimensions['A'].width = 30
            self.ws.column_dimensions['B'].width = 10
            self.ws.column_dimensions['C'].width = 10
            self.ws.column_dimensions['D'].width = 10
            self.ws.column_dimensions['E'].width = 10


class MyMainForm(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.msg_temp = QLabel("")
        self.msg_hum = QLabel("")
        self.msg_lum = QLabel("")
        self.msg_vol = QLabel("")
        self.msg_mag = QLabel("")
        self.lum_thread = None
        self.temp_thread = None
        self.hum_thread = None
        self.vol_thread = None
        self.mag_thread = None
        self.msg_history = list()
        self.setupUi(self)
        self.init_ui()
        self.init_thread()

        # 设置样式
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)  # 设置背景样式：透明
        self.setWindowFlags(Qt.FramelessWindowHint)  # 设置窗口标志：隐藏窗口边框
        # self.lineEdit.setAttribute(QtCore.Qt.WA_MacShowFocusRect, 0)  # ???

    def init_ui(self):
        self.init_log_tem()
        self.init_log_hum()
        self.init_log_lum()
        self.init_log_mag()
        self.init_log_vol()
        # 获取屏幕中心位置坐标
        global SCREEN_X, SCREEN_Y
        center_pointer = QDesktopWidget().availableGeometry().center()
        SCREEN_X = center_pointer.x()
        SCREEN_Y = center_pointer.y()

        # 设置窗口位置,大小参数
        self.move(SCREEN_X - WIN_WIDTH // 2, SCREEN_Y - WIN_HEIGHT // 2)
        self.resize(WIN_WIDTH, WIN_HEIGHT)

        #
        # pic1 = MyPicture

        # 建立信号与槽的连接
        self.listWidget.itemClicked.connect(self.listWidgetFunc)
        self.pushButton_close.clicked.connect(self.click_shift_close)

    def listWidgetFunc(self, item):
        if item.text() == "主 页":
            self.stackedWidget.setCurrentIndex(0)
        elif item.text() == "温 度":
            self.stackedWidget.setCurrentIndex(1)
        elif item.text() == "湿 度":
            self.stackedWidget.setCurrentIndex(2)
        elif item.text() == "光照强度":
            self.stackedWidget.setCurrentIndex(3)
        elif item.text() == "磁场环境":
            self.stackedWidget.setCurrentIndex(4)
        elif item.text() == "环境音量":
            self.stackedWidget.setCurrentIndex(5)

    def click_shift_close(self):
        w_main.close()

    def init_thread(self):
        # 创建线程
        self.temp_thread = TempThread()
        self.hum_thread = HumThread()
        self.lum_thread = LumThread()
        self.vol_thread = VolThread()
        self.mag_thread = MagThread()
        # 开始线程
        self.temp_thread.start()
        self.hum_thread.start()
        self.lum_thread.start()
        self.vol_thread.start()
        self.mag_thread.start()

    def init_log_tem(self):
        self.msg_temp.resize(440, 15)
        self.msg_temp.setWordWrap(True)  # 自动换行 todo
        self.msg_temp.setAlignment(Qt.AlignTop)  # 靠上
        self.scrollArea_2.setWidget(self.msg_temp)
        # 创建垂直布局器，用来添加自动滚动条
        v_layout = QVBoxLayout()
        v_layout.addWidget(self.scrollArea_2)

        # 初始化日志
        msg_history = []
        file = open("tem_log.txt", "r")
        for _ in range(50):
            log = file.readline()
            msg_history.append(log)
            if log == '':
                file.close()
                break

        self.msg_temp.setText("".join(msg_history))
        self.msg_temp.resize(440, self.msg_temp.frameSize().height() + 15)

    def init_log_hum(self):
        self.msg_hum.resize(440, 15)
        self.msg_hum.setWordWrap(True)  # 自动换行 todo
        self.msg_hum.setAlignment(Qt.AlignTop)  # 靠上
        self.scrollArea_3.setWidget(self.msg_hum)
        # 创建垂直布局器，用来添加自动滚动条
        v_layout = QVBoxLayout()
        v_layout.addWidget(self.scrollArea_3)

        # 初始化日志
        msg_history = []
        file = open("hum_log.txt", "r")
        for _ in range(50):
            log = file.readline()
            msg_history.append(log)
            if log == '':
                file.close()
                break

        self.msg_hum.setText("".join(msg_history))
        self.msg_hum.resize(440, self.msg_hum.frameSize().height() + 15)

    def init_log_lum(self):
        self.msg_lum.resize(440, 15)
        self.msg_lum.setWordWrap(True)  # 自动换行 todo
        self.msg_lum.setAlignment(Qt.AlignTop)  # 靠上
        self.scrollArea_4.setWidget(self.msg_lum)
        # 创建垂直布局器，用来添加自动滚动条
        v_layout = QVBoxLayout()
        v_layout.addWidget(self.scrollArea_4)

        # 初始化日志
        msg_history = []
        file = open("lum_log.txt", "r")
        for _ in range(50):
            log = file.readline()
            msg_history.append(log)
            if log == '':
                file.close()
                break

        self.msg_lum.setText("".join(msg_history))
        self.msg_lum.resize(440, self.msg_lum.frameSize().height() + 15)

    def init_log_mag(self):
        self.msg_mag.resize(440, 15)
        self.msg_mag.setWordWrap(True)  # 自动换行 todo
        self.msg_mag.setAlignment(Qt.AlignTop)  # 靠上
        self.scrollArea_5.setWidget(self.msg_mag)
        # 创建垂直布局器，用来添加自动滚动条
        v_layout = QVBoxLayout()
        v_layout.addWidget(self.scrollArea_5)

        # 初始化日志
        msg_history = []
        file = open("mag_log.txt", "r")
        for _ in range(50):
            log = file.readline()
            msg_history.append(log)
            if log == '':
                file.close()
                break

        self.msg_mag.setText("".join(msg_history))
        self.msg_mag.resize(440, self.msg_mag.frameSize().height() + 15)

    def init_log_vol(self):
        self.msg_vol.resize(440, 15)
        self.msg_vol.setWordWrap(True)  # 自动换行 todo
        self.msg_vol.setAlignment(Qt.AlignTop)  # 靠上
        self.scrollArea_6.setWidget(self.msg_vol)
        # 创建垂直布局器，用来添加自动滚动条
        v_layout = QVBoxLayout()
        v_layout.addWidget(self.scrollArea_6)

        # 初始化日志
        msg_history = []
        file = open("vol_log.txt", "r")
        for _ in range(50):
            log = file.readline()
            msg_history.append(log)
            if log == '':
                file.close()
                break

        self.msg_vol.setText("".join(msg_history))
        self.msg_vol.resize(440, self.msg_vol.frameSize().height() + 15)


if __name__ == "__main__":
    app = QApplication(sys.argv)

    w_main = MyMainForm()

    w_main.show()

    sys.exit(app.exec_())